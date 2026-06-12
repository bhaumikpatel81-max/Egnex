"""
KPI Dashboard — visual management dashboard with cards + charts.
Role-scoped: recruiter sees own; ta_manager/admin see all; HM sees their reqs.
Reuses stage_event, sla.py helpers, and existing DB schema — no new migrations.
"""
import io
from datetime import date, timedelta

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from ..db import query, query_one
from ..auth_utils import get_current_user
from ..services.sla import (
    PIPELINE_STAGE_LABELS,
    STAGE_SLA_KEY,
    load_config,
    compute_rag,
)

router = APIRouter(prefix="/api/kpi", tags=["kpi"])

FUNNEL_STAGES = [
    "applied",
    "screen",
    "nexai_bot",
    "shortlisted",
    "interview",
    "documentation",
    "offered",
]

_TERMINAL_STATUSES = (
    "hired", "rejected", "on_hold", "joined",
    "screen_rejected", "dropped", "offer_cancelled",
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _period_start(period: str, year: int) -> date:
    today = date.today()
    p = period.lower()
    if p == "weekly":
        return today - timedelta(days=today.weekday())
    if p == "monthly":
        return date(year, today.month, 1)
    if p == "quarterly":
        m = today.month
        qs = 1 if m <= 3 else 4 if m <= 6 else 7 if m <= 9 else 10
        return date(year, qs, 1)
    if p in ("half_yearly", "half-yearly"):
        return date(year, 4, 1) if 4 <= today.month <= 9 else date(year, 10, 1)
    return date(year, 1, 1)


def _scope(role: str, uid: str):
    """
    Returns (sjoin, swhere, sjp, swp).
    sjp = params consumed by the JOIN condition (before the date param).
    swp = params consumed by the extra WHERE clause (after the date param).
    Pattern in each query: params = sjp + [date_param] + swp
    For queries without a date filter:  params = sjp + swp
    """
    if role == "recruiter":
        return (
            "JOIN requisition_recruiter rr ON rr.requisition_id = r.id AND rr.recruiter_id = %s",
            "",
            [uid],
            [],
        )
    if role == "hiring_manager":
        return ("", "AND r.hiring_manager_id = %s", [], [uid])
    return ("", "", [], [])


# ── Dashboard endpoint ────────────────────────────────────────────────────────

@router.get("/dashboard")
def kpi_dashboard(
    period: str = Query("yearly"),
    year: int = Query(default_factory=lambda: date.today().year),
    user: dict = Depends(get_current_user),
):
    role = user["role"]
    uid  = user["sub"]
    ps   = _period_start(period, year)
    sjoin, swhere, sjp, swp = _scope(role, uid)

    # ── KPI cards ──────────────────────────────────────────────────────────

    open_reqs_row = query_one(
        f"""SELECT COUNT(DISTINCT r.id) AS n
            FROM requisition r {sjoin}
            WHERE r.status = 'open'
              AND COALESCE(r.approval_status, 'approved') = 'approved'
              {swhere}""",
        sjp + swp,
    )

    pos_fill_row = query_one(
        f"""SELECT COALESCE(SUM(r.openings), 0) AS n
            FROM requisition r {sjoin}
            WHERE r.status = 'open'
              AND COALESCE(r.approval_status, 'approved') = 'approved'
              {swhere}""",
        sjp + swp,
    )

    pipeline_row = query_one(
        f"""SELECT COUNT(DISTINCT a.id) AS n
            FROM application a
            JOIN requisition r ON r.id = a.requisition_id {sjoin}
            WHERE a.applied_at >= %s
              AND a.status NOT IN ({', '.join(['%s']*len(_TERMINAL_STATUSES))})
              {swhere}""",
        sjp + [ps] + list(_TERMINAL_STATUSES) + swp,
    )

    ttf_row = query_one(
        f"""SELECT ROUND(AVG(
                EXTRACT(EPOCH FROM (se.occurred_at - a.applied_at)) / 86400.0
            )::numeric, 1) AS avg_days
            FROM application a
            JOIN stage_event se ON se.application_id = a.id AND se.to_status = 'hired'
            JOIN requisition r ON r.id = a.requisition_id {sjoin}
            WHERE se.occurred_at >= %s
            {swhere}""",
        sjp + [ps] + swp,
    )

    # SLA breach counts — reuse sla service helpers
    sla_cfg = load_config()
    breach_rows = query(
        f"""SELECT a.status,
                EXTRACT(EPOCH FROM (now() - COALESCE(
                    (SELECT se2.occurred_at FROM stage_event se2
                     WHERE se2.application_id = a.id
                       AND se2.to_status = a.status
                     ORDER BY se2.occurred_at DESC LIMIT 1),
                    a.applied_at
                ))) / 86400.0 AS elapsed_days
            FROM application a
            JOIN requisition r ON r.id = a.requisition_id {sjoin}
            WHERE a.status NOT IN ({', '.join(['%s']*len(_TERMINAL_STATUSES))})
              AND a.applied_at >= %s
              {swhere}""",
        sjp + list(_TERMINAL_STATUSES) + [ps] + swp,
    )
    red_count = amber_count = 0
    for row in (breach_rows or []):
        sla_key = STAGE_SLA_KEY.get(row["status"], "stage_default")
        target  = sla_cfg.get(sla_key, sla_cfg.get("stage_default", 5))
        rag     = compute_rag(row["elapsed_days"], target)
        if   rag["status"] == "red":   red_count   += 1
        elif rag["status"] == "amber": amber_count += 1

    kpi_cards = {
        "open_reqs":              int(open_reqs_row["n"]) if open_reqs_row else 0,
        "positions_to_fill":      int(pos_fill_row["n"])  if pos_fill_row  else 0,
        "candidates_in_pipeline": int(pipeline_row["n"])  if pipeline_row  else 0,
        "avg_time_to_fill_days": (
            float(ttf_row["avg_days"])
            if ttf_row and ttf_row["avg_days"] is not None
            else None
        ),
        "sla_breaches_red":   red_count,
        "sla_breaches_amber": amber_count,
    }

    # ── Funnel (current status distribution in the period) ─────────────────

    stage_rows = query(
        f"""SELECT a.status, COUNT(*) AS n
            FROM application a
            JOIN requisition r ON r.id = a.requisition_id {sjoin}
            WHERE a.applied_at >= %s
            {swhere}
            GROUP BY a.status""",
        sjp + [ps] + swp,
    )
    stage_map = {r["status"]: int(r["n"]) for r in (stage_rows or [])}

    funnel = []
    prev_count = None
    for s in FUNNEL_STAGES:
        count = stage_map.get(s, 0)
        if s == "offered":
            # candidates who passed through offered (now hired) count here too
            count += stage_map.get("hired", 0) + stage_map.get("joined", 0)
        conv_pct = None
        if prev_count and prev_count > 0:
            conv_pct = round(count / prev_count * 100, 1)
        funnel.append({
            "stage":    s,
            "label":    PIPELINE_STAGE_LABELS.get(s, s.replace("_", " ").title()),
            "count":    count,
            "conv_pct": conv_pct,
        })
        if count > 0:
            prev_count = count

    # ── Source effectiveness ───────────────────────────────────────────────

    source_rows = query(
        f"""SELECT COALESCE(c.source, 'unknown') AS source,
                   COUNT(DISTINCT a.id) AS total,
                   COUNT(DISTINCT a.id) FILTER (
                       WHERE a.status IN ('hired', 'joined')
                   ) AS hires
            FROM application a
            JOIN candidate c ON c.id = a.candidate_id
            JOIN requisition r ON r.id = a.requisition_id {sjoin}
            WHERE a.applied_at >= %s
            {swhere}
            GROUP BY c.source
            ORDER BY total DESC""",
        sjp + [ps] + swp,
    )
    source_effectiveness = [
        {"source": r["source"], "total": int(r["total"]), "hires": int(r["hires"])}
        for r in (source_rows or [])
    ]

    # ── Recruiter load ─────────────────────────────────────────────────────

    if role == "recruiter":
        own = query_one(
            """SELECT
                   COUNT(DISTINCT r.id) FILTER (
                       WHERE r.status = 'open'
                         AND COALESCE(r.approval_status, 'approved') = 'approved'
                   ) AS open_reqs,
                   COUNT(DISTINCT a.id) FILTER (
                       WHERE a.status NOT IN (
                           'hired','rejected','on_hold','joined',
                           'screen_rejected','dropped','offer_cancelled'
                       )
                   ) AS active_candidates
               FROM requisition_recruiter rr
               JOIN requisition r ON r.id = rr.requisition_id
               LEFT JOIN application a ON a.requisition_id = r.id
               WHERE rr.recruiter_id = %s""",
            [uid],
        )
        recruiter_load = [{
            "recruiter":         "You",
            "open_reqs":         int(own["open_reqs"])         if own else 0,
            "active_candidates": int(own["active_candidates"]) if own else 0,
        }]
    else:
        load_rows = query(
            """SELECT u.full_name AS recruiter,
                      COUNT(DISTINCT r.id) FILTER (
                          WHERE r.status = 'open'
                            AND COALESCE(r.approval_status, 'approved') = 'approved'
                      ) AS open_reqs,
                      COUNT(DISTINCT a.id) FILTER (
                          WHERE a.status NOT IN (
                              'hired','rejected','on_hold','joined',
                              'screen_rejected','dropped','offer_cancelled'
                          )
                      ) AS active_candidates
               FROM app_user u
               JOIN requisition_recruiter rr ON rr.recruiter_id = u.id
               JOIN requisition r ON r.id = rr.requisition_id
               LEFT JOIN application a ON a.requisition_id = r.id
               WHERE u.role = 'recruiter'
               GROUP BY u.full_name, u.id
               ORDER BY open_reqs DESC""",
            [],
        )
        recruiter_load = [
            {
                "recruiter":         r["recruiter"],
                "open_reqs":         int(r["open_reqs"] or 0),
                "active_candidates": int(r["active_candidates"] or 0),
            }
            for r in (load_rows or [])
        ]

    # ── Offer stats ────────────────────────────────────────────────────────

    try:
        offer_row = query_one(
            f"""SELECT
                    COUNT(*) FILTER (WHERE o.status IN (
                        'pending_approval','revising','on_hold','draft'
                    )) AS pending,
                    COUNT(*) FILTER (WHERE o.status IN (
                        'approved','sent_to_darwinbox','released','accepted'
                    )) AS approved,
                    COUNT(*) FILTER (WHERE o.status IN (
                        'rejected','cancelled','declined'
                    )) AS rejected,
                    ROUND(AVG(
                        EXTRACT(EPOCH FROM (oas.last_acted - o.created_at)) / 86400.0
                    )::numeric, 1) AS avg_approval_days
                FROM offer o
                JOIN application a ON a.id = o.application_id
                JOIN requisition r ON r.id = a.requisition_id {sjoin}
                LEFT JOIN LATERAL (
                    SELECT MAX(acted_at) AS last_acted
                    FROM offer_approval_step
                    WHERE offer_id = o.id AND acted_at IS NOT NULL
                ) oas ON true
                WHERE o.created_at >= %s
                {swhere}""",
            sjp + [ps] + swp,
        )
        offer_stats = {
            "pending":           int(offer_row["pending"])  if offer_row else 0,
            "approved":          int(offer_row["approved"]) if offer_row else 0,
            "rejected":          int(offer_row["rejected"]) if offer_row else 0,
            "avg_approval_days": (
                float(offer_row["avg_approval_days"])
                if offer_row and offer_row["avg_approval_days"] is not None
                else None
            ),
        }
    except Exception:
        offer_stats = {"pending": 0, "approved": 0, "rejected": 0, "avg_approval_days": None}

    return {
        "kpi_cards":            kpi_cards,
        "funnel":               funnel,
        "source_effectiveness": source_effectiveness,
        "recruiter_load":       recruiter_load,
        "offer_stats":          offer_stats,
    }


# ── Excel export ──────────────────────────────────────────────────────────────

@router.get("/excel")
def kpi_excel(
    period: str = Query("yearly"),
    year: int = Query(default_factory=lambda: date.today().year),
    user: dict = Depends(get_current_user),
):
    import openpyxl
    import openpyxl.styles as xs

    data = kpi_dashboard(period=period, year=year, user=user)

    hdr_fill = xs.PatternFill("solid", fgColor="0C0D10")
    hdr_font = xs.Font(bold=True, color="FFFFFF")

    def _hdr(ws, cols):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(1, ci, col)
            c.fill = hdr_fill
            c.font = hdr_font

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: KPI Summary
    ws1 = wb.create_sheet("KPI Summary")
    _hdr(ws1, ["Metric", "Value"])
    cards = data["kpi_cards"]
    for ri, (k, v) in enumerate([
        ("Open Requisitions",        cards["open_reqs"]),
        ("Positions to Fill",        cards["positions_to_fill"]),
        ("Candidates in Pipeline",   cards["candidates_in_pipeline"]),
        ("Avg Time-to-Fill (days)",  cards["avg_time_to_fill_days"] or "N/A"),
        ("SLA Breaches (Red)",       cards["sla_breaches_red"]),
        ("SLA Warnings (Amber)",     cards["sla_breaches_amber"]),
    ], 2):
        ws1.cell(ri, 1, k)
        ws1.cell(ri, 2, v)

    # Sheet 2: Pipeline Funnel
    ws2 = wb.create_sheet("Pipeline Funnel")
    _hdr(ws2, ["Stage", "Candidates", "Conversion %"])
    for ri, f in enumerate(data["funnel"], 2):
        ws2.cell(ri, 1, f["label"])
        ws2.cell(ri, 2, f["count"])
        ws2.cell(ri, 3, f["conv_pct"] if f["conv_pct"] is not None else "")

    # Sheet 3: Source Effectiveness
    ws3 = wb.create_sheet("Source Effectiveness")
    _hdr(ws3, ["Source", "Total Candidates", "Hires", "Hit Rate %"])
    for ri, s in enumerate(data["source_effectiveness"], 2):
        ws3.cell(ri, 1, s["source"])
        ws3.cell(ri, 2, s["total"])
        ws3.cell(ri, 3, s["hires"])
        ws3.cell(ri, 4, round(s["hires"] / s["total"] * 100, 1) if s["total"] else 0)

    # Sheet 4: Recruiter Load
    ws4 = wb.create_sheet("Recruiter Load")
    _hdr(ws4, ["Recruiter", "Open Reqs", "Active Candidates"])
    for ri, r in enumerate(data["recruiter_load"], 2):
        ws4.cell(ri, 1, r["recruiter"])
        ws4.cell(ri, 2, r["open_reqs"])
        ws4.cell(ri, 3, r["active_candidates"])

    # Sheet 5: Offer Stats
    ws5 = wb.create_sheet("Offer Stats")
    _hdr(ws5, ["Metric", "Value"])
    o = data["offer_stats"]
    for ri, (k, v) in enumerate([
        ("Pending",           o["pending"]),
        ("Approved",          o["approved"]),
        ("Rejected",          o["rejected"]),
        ("Avg Approval Days", o["avg_approval_days"] or "N/A"),
    ], 2):
        ws5.cell(ri, 1, k)
        ws5.cell(ri, 2, v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=egnex_kpi_{year}_{period}.xlsx"
        },
    )
